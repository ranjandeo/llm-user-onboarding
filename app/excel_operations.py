import openpyxl
import logging
import json
from app.customer import create_new_customer
from app.key_management import generate_key
from utils.validators import is_valid_email


def load_excel(file_name, sheet_name):
    """
    Load an Excel workbook and select a specified sheet.
    """
    try:
        wb = openpyxl.load_workbook(filename=file_name)
        sheet = wb[sheet_name]
        logging.info(
            f'Excel file "{file_name}" loaded and sheet "{sheet_name}" selected successfully.'
        )
        return wb, sheet
    except Exception as e:
        logging.error(f"Failed to load Excel file or select sheet: {str(e)}")
        raise


def process_emails(wb, sheet, file_name, api_key):
    """
    Process each email in the Excel sheet, creating customers and generating keys.
    """
    invalid_emails = []

    try:
        for row in sheet.iter_rows(min_row=2, max_col=2, values_only=False):
            email_cell = row[0]
            email = (
                email_cell.value.lower() if email_cell.value else None
            )  # Convert email to lowercase
            if email:
                logging.debug(f"Processing email: {email}")

                if not is_valid_email(email):
                    invalid_emails.append(email)
                    logging.warning(f"Invalid email format: {email}")
                    continue

                alias = email.split("@")[0]

                result_new_customer = create_new_customer(email, alias, api_key)
                result_generate_key = generate_key(email, alias, api_key)

                try:
                    key_response = json.loads(result_generate_key)
                    generated_key = key_response.get("key", "")

                    key_cell = row[1]
                    key_cell.value = generated_key
                except json.JSONDecodeError:
                    logging.error(
                        "Failed to parse JSON response from generate key command"
                    )
            else:
                logging.warning(f"No email found in row: {row}")

        wb.save(file_name)
        logging.info("Changes saved to the Excel file successfully")

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")

    return invalid_emails


def save_invalid_emails(wb, invalid_emails, file_name):
    """
    Save changes to the Excel workbook and log any invalid emails.
    """

    try:
        # Add invalid emails to a new sheet
        if invalid_emails:
            if "Invalid Emails" in wb.sheetnames:
                invalid_sheet = wb["Invalid Emails"]

                # Clear existing data
                invalid_sheet.delete_rows(2, invalid_sheet.max_row - 1)
            else:
                invalid_sheet = wb.create_sheet(title="Invalid Emails")
                invalid_sheet.append(["Invalid Emails"])

            logging.info("The below email addresse(s) are invalid:")

            for invalid_email in invalid_emails:
                invalid_sheet.append([invalid_email])
                logging.info(invalid_email)

            wb.save(file_name)
            logging.info("Invalid emails written to the new sheet successfully")

    except Exception as e:
        logging.error(f"Error saving Excel file: {str(e)}")
        raise
